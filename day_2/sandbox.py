import os
import openpyxl

<<<<<<< HEAD
wb = openpyxl.load_workbook('day_2/lecture.xlsx')
# workbook = openpyxl.load_workbook('day_2/lecture.xlsx')
# print(wb.sheetnames)

# wb = Workbook()
# print(str(wb))

# new_sheet = wb.create_sheet()
# print(str(new_sheet))

for i in range(3, 10):
    wb.create_sheet()

for sheet in wb:
    y = int(50)
    sheet.title = "MyNewSheets" + y
    y += 1

print(str(wb.sheetnames))
=======
workbook = openpyxl.load_workbook('day_2/lecture.xlsx')
print(str(workbook))
>>>>>>> 5316a6a19ffbb5e750b61f0a1628d822854fd69b
